import json
import os

import pandas as pd
from datasets import Dataset
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

from opencompass.openicl.icl_evaluator import BaseEvaluator
from opencompass.registry import LOAD_DATASET

from ..base import BaseDataset
# TODO: use ifeval config
from .evaluation_main import InputExample, test_instruction_following_strict

instructs = [
    'detectable_format:multiple_sections',
    'detectable_format:number_bullet_lists',
    'length_constraints:number_words',
    'length_constraints:number_sentences',
    'length_constraints:nth_paragraph_first_word',
    'startend:end_checker',
    'startend:quotation',
    'detectable_format:number_highlighted_sections',
    'punctuation:no_punctuation',
    'combination:two_responses',
    'language:forbidden_language',
    'keywords:forbidden_words',
    'keywords:existence',
    'llm_as_judge',
    'detectable_content:constrain_choices',
    'length_constraints:summary_length',
    'combination:repeat_prompt',
]


@LOAD_DATASET.register_module()
class FollowLawDataset(BaseDataset):

    @staticmethod
    def load(path, name=None):
        # path = get_data_path(path, local_mode=True)
        datasets = []
        with open(os.path.join(path, f'{name}'), 'r',
                  encoding='utf-8') as file:
            for line in file:
                tmp = json.loads(line.strip())
                dataset = dict(prompt=tmp['prompt'], reference=tmp)
                datasets.append(dataset)
        return Dataset.from_list(datasets)


class FollowLawEvaluator(BaseEvaluator):

    def score(self, predictions, references, origin_prompt):
        prompt_strict_correct, prompt_strict_total = 0, 0
        inst_strict_correct, inst_strict_total = 0, 0
        # prompt_loose_correct, prompt_loose_total = 0, 0
        # inst_loose_correct, inst_loose_total = 0, 0

        details = {}
        for index, (pred, refer) in enumerate(zip(predictions, references)):
            input = InputExample(
                key=refer['key'],
                instruction_id_list=refer['instruction_id_list'],
                prompt=refer['prompt'],
                requirements=refer['requirements'],
                kwargs=refer['kwargs'])
            for kwarg in input.kwargs:
                for k in list(kwarg.keys()):
                    if kwarg[k] is None:
                        kwarg.pop(k, None)

            # strict
            example = test_instruction_following_strict(input, pred)
            reasons = example.reasons
            follow_instruction_list = example.follow_instruction_list
            instruction_id_list = example.instruction_id_list
            prompt_strict_total += 1
            is_strict_correct = all(follow_instruction_list)
            prompt_strict_correct += is_strict_correct
            inst_strict_total += len(instruction_id_list)
            inst_strict_correct += sum(follow_instruction_list)

            if is_strict_correct:
                grade = 'strict'
            # elif is_loose_correct:
            #     grade = 'loose'
            else:
                grade = 'none'

            details[str(index)] = {
                'prompt': origin_prompt[index],
                'pred': pred,
                'refer': refer,
                'follow_list': follow_instruction_list,
                'is_strict_correct': is_strict_correct,
                # 'is_loose_correct': is_loose_correct,
                'reasons': reasons,
                'is_correct': is_strict_correct,
                'grade': grade
            }

        results = {
            'Prompt-level-strict-accuracy':
            prompt_strict_correct / prompt_strict_total * 100,
            'Inst-level-strict-accuracy':
            inst_strict_correct / inst_strict_total * 100,

            # 'Prompt-level-loose-accuracy':
            # prompt_loose_correct / prompt_loose_total * 100,
            # 'Inst-level-loose-accuracy':
            # inst_loose_correct / inst_loose_total * 100,
            'details':
            details
        }

        all_infos = []
        for qid, line in details.items():
            all_infos.append({
                'key_id':
                line['refer']['key'],
                'prompt':
                line['prompt'],
                'answer':
                line['pred'],
                'instruction_id_list':
                num_bullet_list(line['refer']['instruction_id_list']),
                'requirements':
                num_bullet_list(line['refer']['requirements']),
                'kwargs':
                num_bullet_list(line['refer']['kwargs']),
                'follow_list':
                num_bullet_list(line['follow_list']),
                # 'is_strict_correct': line["is_strict_correct"],
                # 'is_loose_correct': line["is_loose_correct"],
                'is_correct':
                line['is_correct'],
                'reasons':
                num_bullet_list(line['reasons']),
                'grade':
                line['grade']
            })
        # with open("_debug.") as f:

        df = pd.DataFrame(all_infos)
        dataset_name = os.environ.get('dataset_name', '')
        os.makedirs('outputs/followlaw', exist_ok=True)
        file_name = f'outputs/followlaw/{dataset_name}-follow_debug.xlsx'
        df.to_excel(file_name, index=False)  # index=False 用于不保存行索引

        # 使用 openpyxl 调整列宽和设置自动换行
        wb = load_workbook(file_name)
        ws = wb.active

        # 设置列宽和自动换行
        for col_num, col_cells in enumerate(ws.columns, 1):  # 遍历每一列
            max_length = 0
            for cell in col_cells:
                try:
                    if cell.value:
                        max_length = max(max_length,
                                         min(len(str(cell.value)), 30))
                except Exception as e:
                    print(e)
            col_letter = get_column_letter(col_num)  # 获取列字母
            ws.column_dimensions[col_letter].width = max_length + 2  # 设置列宽
            for cell in col_cells:
                cell.alignment = Alignment(wrap_text=True)  # 设置自动换行

        # 保存格式化后的文件
        wb.save(file_name)
        return results


def num_bullet_list(elements):
    return '\n'.join(
        [f'{i + 1}. {element}' for i, element in enumerate(elements)])
